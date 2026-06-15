#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""monoqrome SNS フォロワー数 デイリー記録（Claude Code ルーティーン版）

Claude デスクトップ版（Claude in Chrome によるブラウザ自動操作）から移植。
ブラウザ自動操作を一切使わず、curl + 正規表現で公開数値を取得し、
リポジトリ内 Excel に追記、ログに一次記録、git コミット/プッシュ（ベストエフォート）
までを行う完全自律スクリプト。launchd から無人実行される前提。

設計方針（元タスク準拠）:
- 完全自律       : 確認・許可待ちは一切しない。
- 途中停止禁止   : どのステップが失敗しても中断せず次へ。値が取れなければ "取得失敗"。
- 一次記録を最優先: 数値は Excel 保存より先にログへ書き込む。Excel が失敗しても
                    ログを遡れば値が分かるようにする（元タスクの「応答冒頭ブロック」相当）。

実行:  /usr/bin/python3 scripts/monoqrome_sns_tracker.py
"""
import os
import re
import sys
import subprocess
import datetime

# launchd の最小 PATH でも curl / git を解決できるよう明示的に補完
os.environ["PATH"] = "/usr/bin:/bin:/usr/sbin:/sbin:/usr/local/bin:/opt/homebrew/bin:" + os.environ.get("PATH", "")

REPO  = "/Users/dz0019/Desktop/claude_management"
XLSX  = os.path.join(REPO, "03_projects/MONOQROME/monoqrome_sns_tracker.xlsx")
LOG   = os.path.join(REPO, "scripts/monoqrome_sns_tracker.log")
SHEET = "フォロワー記録"

YT_URL = "https://www.youtube.com/@monoqrome"
IG_URL = "https://www.instagram.com/monoqrome_info/"

UA_DESKTOP = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
              "(KHTML, like Gecko) Chrome/120.0 Safari/537.36")
UA_MOBILE  = ("Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) "
              "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1")


# ---------------------------------------------------------------------------
# 取得
# ---------------------------------------------------------------------------
def curl(url, ua, cookie=None, timeout=25):
    cmd = ["/usr/bin/curl", "-sL", "--max-time", str(timeout), "-A", ua]
    if cookie:
        cmd += ["-H", f"Cookie: {cookie}"]
    cmd.append(url)
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout + 5)
        return r.stdout or ""
    except Exception as e:
        print(f"curl error ({url}): {e}", file=sys.stderr)
        return ""


def get_youtube():
    """YouTube 登録者数。consent cookie 付き curl で HTML を取得し数値を抽出。"""
    html = curl(YT_URL, UA_DESKTOP, cookie="SOCS=CAI; CONSENT=YES+1")
    if not html:
        return None
    # 1) 表示テキスト "登録者数 5,120人" / "チャンネル登録者数 5120人"
    m = re.search(r"登録者数[^\d]{0,4}([\d,]+)\s*人", html)
    if m:
        return int(m.group(1).replace(",", ""))
    # 2) JSON subscriberCountText 内の数字
    m = re.search(r'"subscriberCountText".{0,160}?([\d,]+)\s*人', html)
    if m:
        return int(m.group(1).replace(",", ""))
    # 3) 1万人超で万表記になった場合（精度は落ちるが取得失敗を避ける）
    m = re.search(r"登録者数[^\d]{0,4}([\d.]+)\s*万\s*人", html)
    if m:
        return int(round(float(m.group(1)) * 10000))
    return None


def get_instagram():
    """Instagram フォロワー数。og:description の "X Followers" を抽出。"""
    for ua in (UA_MOBILE, UA_DESKTOP):
        html = curl(IG_URL, ua)
        if not html:
            continue
        # og:description: 'X Followers, Y Following, Z Posts - ...'
        m = re.search(r'content="([\d,]+)\s*Followers', html)
        if m:
            return int(m.group(1).replace(",", ""))
        m = re.search(r'([\d,]+)\s*Followers', html)
        if m:
            return int(m.group(1).replace(",", ""))
    return None


# ---------------------------------------------------------------------------
# Excel 保存
# ---------------------------------------------------------------------------
def update_excel(ig, yt, today):
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    ig_val = ig if isinstance(ig, int) else "取得失敗"
    yt_val = yt if isinstance(yt, int) else "取得失敗"

    if os.path.exists(XLSX):
        wb = load_workbook(XLSX)
        ws = wb[SHEET] if SHEET in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET
        ws.append(["日付", "Instagram フォロワー数", "YouTube 登録者数",
                   "Instagram 前日比", "YouTube 前日比"])
        hf = Font(bold=True, color="FFFFFF")
        hfill = PatternFill("solid", start_color="2E4057")
        for c in range(1, 6):
            ws.cell(row=1, column=c).font = hf
            ws.cell(row=1, column=c).fill = hfill
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for w, col in zip([15, 25, 22, 20, 18], "ABCDE"):
            ws.column_dimensions[col].width = w

    # 同日行があれば上書き、なければ追記
    target = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value) == today:
            target = r
            break
    if target is None:
        target = ws.max_row + 1
        ws.cell(row=target, column=1, value=today)

    ws.cell(row=target, column=2, value=ig_val)
    ws.cell(row=target, column=3, value=yt_val)

    prev_ig = ws.cell(row=target - 1, column=2).value if target >= 3 else None
    prev_yt = ws.cell(row=target - 1, column=3).value if target >= 3 else None
    if target >= 3 and isinstance(ig_val, (int, float)) and isinstance(prev_ig, (int, float)):
        c = ws.cell(row=target, column=4, value=f"=B{target}-B{target-1}")
        c.number_format = "+#,##0;-#,##0;0"
    if target >= 3 and isinstance(yt_val, (int, float)) and isinstance(prev_yt, (int, float)):
        c = ws.cell(row=target, column=5, value=f"=C{target}-C{target-1}")
        c.number_format = "+#,##0;-#,##0;0"

    wb.save(XLSX)
    return target


# ---------------------------------------------------------------------------
# ログ（一次記録）
# ---------------------------------------------------------------------------
def fmt(v):
    return str(v) if isinstance(v, int) else "取得失敗"


def write_log(text):
    print(text, end="")
    try:
        with open(LOG, "a", encoding="utf-8") as f:
            f.write(text)
    except Exception as e:
        print(f"log write failed: {e}", file=sys.stderr)


# ---------------------------------------------------------------------------
# git 記録（ベストエフォート・絶対にハングさせない）
# ---------------------------------------------------------------------------
def git_record(today):
    env = {**os.environ, "GIT_TERMINAL_PROMPT": "0"}

    def run(args, timeout=30):
        return subprocess.run(["/usr/bin/git", "-C", REPO] + args,
                              capture_output=True, text=True, timeout=timeout, env=env)
    try:
        # 追跡対象を本タスクの2ファイルだけに限定（ユーザーの作業中変更を巻き込まない）
        run(["add", XLSX, LOG])
        if run(["diff", "--cached", "--quiet"]).returncode == 0:
            return "コミット差分なし"
        run(["commit", "-m", f"monoqrome SNS記録 {today}"])
        push = run(["push", "origin", "main"], timeout=60)
        return "コミット+プッシュ成功" if push.returncode == 0 else "コミットのみ（push失敗）"
    except Exception as e:
        return f"スキップ（{e}）"


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main():
    today = datetime.date.today().strftime("%Y-%m-%d")
    now   = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # 1) 取得
    ig = get_instagram()
    yt = get_youtube()

    # 2) 一次記録（数値）を Excel より先に必ず残す
    write_log(
        f"[{now}] 【monoqrome SNS 記録 {today}】\n"
        f"- Instagram (@monoqrome_info): {fmt(ig)}\n"
        f"- YouTube (@monoqrome): {fmt(yt)}\n"
    )

    # 3) Excel 保存（失敗しても中断しない）
    try:
        row = update_excel(ig, yt, today)
        excel_status = f"成功 (row={row})"
    except Exception as e:
        excel_status = f"失敗（{e}）"

    # 4) git 記録（ベストエフォート）
    try:
        git_status = git_record(today)
    except Exception as e:
        git_status = f"失敗（{e}）"

    # 5) 結果ログ
    write_log(f"  Excel保存: {excel_status} / git: {git_status}\n\n")


if __name__ == "__main__":
    main()